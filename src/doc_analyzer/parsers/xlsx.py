"""Excel (.xlsx) parsing via openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from doc_analyzer.models import Document

logger = logging.getLogger(__name__)


class XlsxParserError(Exception):
    """Base class for XLSX parsing failures."""


class XlsxFileNotFoundError(XlsxParserError):
    """The path does not exist or is not a regular file."""


class XlsxMalformedError(XlsxParserError):
    """openpyxl cannot parse the file as a valid .xlsx."""


def parse_xlsx(path: str | Path) -> Document:
    """Parse an Excel (.xlsx) file into a :class:`Document`.

    Each worksheet becomes one page. Cells are joined tab-separated per
    row; empty rows are dropped. Formula cells are evaluated to their
    cached value via ``data_only=True``.
    """
    xlsx_path = Path(path)
    if not xlsx_path.is_file():
        raise XlsxFileNotFoundError(f"XLSX file does not exist: {xlsx_path}")

    try:
        workbook = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    except (InvalidFileException, KeyError) as exc:
        raise XlsxMalformedError(f"Failed to open XLSX: {xlsx_path}") from exc
    except Exception as exc:
        raise XlsxMalformedError(f"Unexpected error reading XLSX: {xlsx_path}: {exc}") from exc

    pages: list[str] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_text: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cell.strip() for cell in cells):
                    rows_text.append("\t".join(cells))
            page = f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text)
            pages.append(page)
    finally:
        workbook.close()

    if not pages:
        pages = [""]

    return Document(source=xlsx_path, pages=tuple(pages))
